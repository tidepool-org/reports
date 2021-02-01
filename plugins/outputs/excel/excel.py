"""
Copyright (c) 2020, Tidepool Project
All rights reserved.
"""
import logging
from typing import Tuple, List
from operator import attrgetter
import openpyxl
import re
import ast

from .html import HtmlToExcel
from .column import Column
from .columns import Columns
import plugins.output
from ...inputs.jira import JiraRiskScore

logger = logging.getLogger(__name__)

def log_issue(issue, indent: int = 0):
    logger.debug(f"{indent * '--> '}added {issue.type} {issue.key} {issue.url}")

class Excel(plugins.output.OutputGenerator):
    key = 'excel'
    flag = '--excel'
    description = 'generate Excel output'
    _alias_ = 'Excel'

    @property
    def labels(self):
        return self.config['labels']

    @property
    def generated_date(self):
        return self.config['generated'].astimezone().strftime('%Y-%m-%d %H:%M:%S %Z')

    @property
    def build_number(self):
        return self.config['build']

    @property
    def tag(self):
        return self.config['tag']

    @property
    def passed(self):
        return self.labels['passed']

    @property
    def blocked(self):
        return self.labels['blocked']

    @property
    def verified(self):
        return self.labels['verified']

    def generate(self) -> List[str]:
        template_file = self.config['template']['report']
        output_file = self.config['output']['report']
        logger.info(f"generating {output_file} from {template_file}")

        book = openpyxl.load_workbook(template_file)

        for format_key, format in self.config['formats'].items():
            book.add_named_style(self.create_format(format_key, format))

        self.risk_formats = {
            JiraRiskScore.GREEN: 'low_risk',
            JiraRiskScore.YELLOW: 'medium_risk',
            JiraRiskScore.RED: 'high_risk',
            JiraRiskScore.UNKNOWN: 'unknown_risk',
        }

        for sheet in book.worksheets:
            logger.info(f"examining sheet '{sheet.title}'")
            for row in sheet.iter_rows():
                for cell in row:
                    # substitution?
                    if re.search(r'\{.+\}', str(cell.value)):
                        text = self.format_text(cell.value)
                        logger.info(f"replacing '{cell.value}' with '{text}'")
                        cell.value = text
                    # insertion?
                    insertion = re.search(r'<<<insert: (\w+)\((.*)\)>>>', str(cell.value))
                    if insertion:
                        generator_method = getattr(self.__class__, insertion.group(1))
                        props = { }
                        if insertion.group(2):
                            logger.debug(f"parsing {insertion.group(2)}")
                            props = ast.literal_eval('{' + insertion.group(2) + '}')
                            logger.debug(f"got {props}")
                        generator_method(self, sheet, cell.row, cell.column, props = props)
                    else:
                        self.set_paper(sheet)

        book.save(output_file)

        logger.info(f"done generating {output_file}")
        return [ output_file ]

    #
    # Software Requirements
    #
    def requirements_list(self, sheet: openpyxl.worksheet, start_row: int, start_col: int, props: dict = { }) -> None:
        logger.info(f"adding report sheet '{sheet.title}'")

        # requirements, sorted by requirement ID
        req_ids = { }
        row = start_row
        for req in self.jira.sorted_by_id(self.jira.exclude_junk(self.jira.func_requirements.values(), enforce_versions = False)):
            log_issue(req)

            col = start_col
            self.write_id(sheet, row, col, req)
            self.write_key(sheet, row, col + 1, req)
            self.write(sheet, row, col + 2, req.summary)
            self.write_html(sheet, row, col + 3, req.description)

            row += 1
            # uniqueness check; requirement IDs should be globally unique
            if req.id in req_ids:
                logger.warn(f"requirement ID '{req.id}' duplicated in: {', '.join([ *req_ids[req.id], req.key ])}")
                req_ids[req.id].append(req.key)
            else:
                req_ids[req.id] = [ req.key ]

        # report.ignore_errors({ 'number_stored_as_text': xl_range(0, 0, row - 1, columns.last) })
        self.set_paper(sheet, start_row - 1)
        logger.info(f'total of {len(req_ids)} requirements')
        dupes = { req_id: req_keys for req_id, req_keys in req_ids.items() if len(req_keys) > 1 }
        if len(dupes) > 0:
            logger.warn(f"total of {len(dupes)} duplicated requirement IDs: {', '.join(dupes.keys())}")
        logger.info(f"done adding report sheet '{sheet.title}'")

    #
    # Traceability Report (full and not-so-full)
    #
    def traceability_report(self, sheet: openpyxl.worksheet, start_row: int, start_col: int, props: dict = { }) -> None:
        logger.info(f"adding report sheet '{sheet.title}'")

        # requirements, sorted by requirement ID
        total_requirements = 0
        total_verified = 0
        total_future = 0
        row = start_row
        if 'filter' in props:
            logger.info(f"filtering requirements by keys: {props['filter']}")
        if 'filter_id' in props:
            logger.info(f"filtering requirements by ID: {props['filter_id']}")
        for req in self.jira.sorted_by_id(self.jira.exclude_junk(self.jira.filter_by_id(self.jira.filter_by_key(self.jira.func_requirements.values(), props.get('filter')), props.get('filter_id')), enforce_versions = False)):
            log_issue(req)
            req_row = row
            col = start_col

            # stories, sorted by issue key
            verified = False
            story_row = req_row
            story_col = col + 4
            stories = self.jira.sorted_by_key(self.jira.exclude_junk(req.defines, enforce_versions = True))
            for story in stories:
                log_issue(story, 1)

                # tests, sorted by issue key
                test_row, tests_verified = self.write_tests(sheet, story_row, story_col + 2, story)

                # update verification status
                verified = verified or story.is_done or tests_verified

                # story summary, possibly across many rows
                row = max(story_row + 1, test_row) - 1
                self.write_key_and_summary(sheet, story_row, story_col, story, end_row = row)
                self.set_outline(sheet, story_row, req_row, 1)
                story_row = row + 1

            if req.is_device_qualification and story_row == req_row: # there were no attached implementation stories?
                logger.info(f"{req.key} {req.id} '{req.summary}' is a device qualification requirement -> future verification")
                self.write(sheet, story_row, story_col, self.labels['device_qual_req'], format = 'bold', end_col = story_col + 4)
                total_future += 1
            else:
                if verified:
                    total_verified += 1
                else:
                    if len(stories) > 0:
                        logger.warn(f"{req.key} {req.id} '{req.summary}' has {len(stories)} linked stories that implement it, but none verify it")
                    else:
                        logger.warn(f"{req.key} {req.id} '{req.summary}' has no linked stories that implement it")

            risk_row = req_row
            if props.get('full'): # include risks?
                risk_row, mitigated = self.write_risks(sheet, risk_row, story_col + 6, req, filter = props.get('filter_risks'))

            row = max(req_row + 1, risk_row, story_row) - 1
            self.write_id(sheet, req_row, col, req, end_row = row)
            self.write_key_and_summary(sheet, req_row, col + 1, req, end_row = row)
            self.write_html(sheet, req_row, col + 3, req.description, end_row = row)
            if props.get('full'):
                self.merge(sheet, req_row, story_col + 5, end_row = row)
            row += 1
            total_requirements += 1

        # report.ignore_errors({ 'number_stored_as_text': xl_range(0, 0, row - 1, columns.last) })
        self.set_paper(sheet, start_row - 1)
        logger.info(f'total of {total_requirements} requirements')
        logger.info(f'total of {self.percentage(total_requirements, total_verified)} verified requirements')
        logger.info(f'total of {self.percentage(total_requirements, total_future)} requirements to be verified in the future')
        logger.info(f"done adding report sheet '{sheet.title}'")

    #
    # Traceability Summary
    #
    def traceability_summary(self, sheet: openpyxl.worksheet, start_row: int, start_col: int, props: dict = { }) -> None:
        logger.info(f"adding report sheet '{sheet.title}'")

        # requirements, sorted by requirement ID
        total_requirements = 0
        row = start_row
        for req in self.jira.sorted_by_id(self.jira.exclude_junk(self.jira.func_requirements.values(), enforce_versions = False)):
            log_issue(req)
            req_row = row
            col = start_col

            # stories, sorted by issue key
            story_row = req_row
            story_col = col + 4
            for story in self.jira.sorted_by_key(self.jira.exclude_junk(req.defines, enforce_versions = True)):
                log_issue(story, 1)
                self.write_key_and_summary(sheet, story_row, story_col, story)
                if story.is_done:
                    self.write(sheet, story_row, story_col + 2, self.verified, format = 'bold')
                self.set_outline(sheet, story_row, req_row, 1)
                story_row += 1

            if req.is_device_qualification and story_row == req_row: # there were no attached implementation stories?
                logger.info(f"{req.key} {req.id} '{req.summary}' is a device qualification requirement -> future verification")
                self.write(sheet, story_row, story_col, self.labels['device_qual_req'], format = 'bold', end_col = story_col + 2)

            row = max(req_row + 1, story_row) - 1
            self.write_id(sheet, req_row, col, req, end_row = row)
            self.write_key_and_summary(sheet, req_row, col + 1, req, end_row = row)
            self.write_html(sheet, req_row, col + 3, req.description, end_row = row)
            row += 1
            total_requirements += 1

        # report.ignore_errors({ 'number_stored_as_text': xl_range(0, 0, row - 1, columns.last) })
        self.set_paper(sheet, start_row - 1)
        logger.info(f'total of {total_requirements} requirements')
        logger.info(f"done adding report sheet '{sheet.title}'")

    #
    # Verification Test Report
    #
    def verification_report(self, sheet: openpyxl.worksheet, start_row: int, start_col: int, props: dict = { }) -> None:
        logger.info(f"adding report sheet '{sheet.title}'")

        # stories, sorted by issue key
        row = start_row
        col = start_col
        stories = self.jira.sorted_by_key(self.jira.exclude_junk(self.jira.stories.values(), enforce_versions = True))
        for story in stories:
            log_issue(story, 1)
            story_row = row

            # tests, sorted by issue key
            test_row, tests_verified = self.write_tests(sheet, story_row, col + 2, story)

            # update verification status
            verified = story.is_done or tests_verified

            # story summary, possibly across many rows
            row = max(story_row + 1, test_row) - 1
            self.write_key_and_summary(sheet, story_row, col, story, end_row = row)
            self.set_outline(sheet, story_row, row, 1)

            row += 1

        logger.info(f"done adding report sheet '{sheet.title}'")

    #
    # Hazard Analysis
    #
    def hazard_analysis(self, sheet: openpyxl.worksheet, start_row: int, start_col: int, props: dict = { }) -> None:
        logger.info(f"adding report sheet '{sheet.title}'")

        # risks, sorted by harm
        total_risks = 0
        total_initial_scores = self.jira.risk_scores
        total_residual_scores = self.jira.risk_scores
        row = start_row
        if 'filter' in props:
            logger.info(f"filtering risks by {props['filter']}")
        for risk in self.jira.sorted_by_harm(self.jira.exclude_junk(self.jira.filter_by_key(self.jira.risks.values(), props.get('filter')), enforce_versions = False)):
            log_issue(risk)
            risk_row = row
            col = start_col
            total_risks += 1
            initial_risk_score = risk.score(risk.initial_risk, 'initial')
            total_initial_scores[initial_risk_score] += 1
            residual_risk_score = risk.score(risk.residual_risk, 'residual')
            total_residual_scores[residual_risk_score] += 1

            # list all mitigations in the sheet
            story_row = row
            story_col = col + 10
            offset = 0
            if 'mitigation_id' in props:
                offset = 1
            for mitigation in self.jira.sorted_by_key(risk.mitigations):
                if offset > 0 and mitigation.is_func_requirement:
                    self.write(sheet, story_row, story_col + 0, mitigation.id)
                self.write_key_and_summary(sheet, story_row, story_col + offset + 0, mitigation)
                logger.debug(f"""{mitigation.key}: '{mitigation.description}'""")
                self.write_html(sheet, story_row, story_col + offset + 2, mitigation.description)
                self.set_outline(sheet, story_row, row, 1)
                story_row += 1

            row = max(risk_row + 1, story_row) - 1
            self.write_key_and_summary(sheet, risk_row, col, risk, end_row = row)
            self.write(sheet, risk_row, col + 2, risk.source, end_row = row)
            self.write_html(sheet, risk_row, col + 3, risk.sequence, end_row = row)
            self.write(sheet, risk_row, col + 4, risk.harm, end_row = row)
            self.write(sheet, risk_row, col + 5, risk.hazard, end_row = row)
            self.write(sheet, risk_row, col + 6, risk.hazard_category, end_row = row)
            self.write(sheet, risk_row, col + 7, risk.initial_severity, end_row = row)
            self.write(sheet, risk_row, col + 8, risk.initial_probability, end_row = row)
            self.write(sheet, risk_row, col + 9, risk.initial_risk, format = self.risk_formats[initial_risk_score], end_row = row)
            self.write(sheet, risk_row, col + offset + 13, risk.residual_severity, end_row = row)
            self.write(sheet, risk_row, col + offset + 14, risk.residual_probability, end_row = row)
            self.write(sheet, risk_row, col + offset + 15, risk.residual_risk, format = self.risk_formats[residual_risk_score], end_row = row)
            row += 1

        self.set_paper(sheet, start_row - 1)
        self.write_risk_summary(sheet, 1, total_risks, total_initial_scores, col + 9, total_residual_scores, col + 15)
        logger.info(f"done adding report sheet '{sheet.title}'")

    #
    # insulin fidelity
    #
    def insulin_fidelity(self, sheet: openpyxl.worksheet, start_row: int, start_col: int, props: dict = { }) -> None:
        logger.info(f"adding report sheet '{sheet.title}'")

        # risks, sorted by harm
        total_risks = 0
        total_initial_scores = self.jira.risk_scores
        total_residual_scores = self.jira.risk_scores
        row = start_row
        for risk in self.jira.sorted_by_harm(self.jira.filter_by_key(self.jira.risks.values(), props.get('filter'))):
            log_issue(risk)
            risk_row = row
            col = start_col
            total_risks += 1
            initial_risk_score = risk.score(risk.initial_risk, 'initial')
            total_initial_scores[initial_risk_score] += 1
            residual_risk_score = risk.score(risk.residual_risk, 'residual')
            total_residual_scores[residual_risk_score] += 1

            # list all mitigations in the sheet
            story_row = row
            story_col = col + 7
            for mitigation in self.jira.sorted_by_key(risk.mitigations):
                stories = set()
                tests = set()
                if mitigation.is_func_requirement:
                    self.write_id(sheet, story_row, story_col, mitigation)
                    self.write_html(sheet, story_row, story_col + 1, mitigation.description)
                    for story in mitigation.defines: # add all stories, and all tests that verify those stories
                        stories.add(story)
                        tests.update(story.tests)
                else: # story or IFU
                    stories.add(mitigation)
                    tests.update(mitigation.tests)
                if len(tests) == 0: # if there are no tests, then the stories cover the testing
                    tests = stories
                self.write(sheet, story_row, story_col + 2, ', '.join([ story.key for story in self.jira.sorted_by_key(stories) ]))
                self.write(sheet, story_row, story_col + 3, ', '.join([ test.key for test in self.jira.sorted_by_key(tests) ]))
                self.set_outline(sheet, story_row, row, 1)
                story_row += 1

            row = max(risk_row + 1, story_row) - 1
            self.write_key_and_summary(sheet, risk_row, col, risk, end_row = row)
            self.write(sheet, risk_row, col + 2, risk.hazard, end_row = row)
            self.write(sheet, risk_row, col + 3, risk.hazard_category, end_row = row)
            self.write(sheet, risk_row, col + 4, risk.initial_severity, end_row = row)
            self.write(sheet, risk_row, col + 5, risk.initial_risk, format = self.risk_formats[initial_risk_score], end_row = row)
            self.write(sheet, risk_row, col + 6, risk.residual_risk, format = self.risk_formats[residual_risk_score], end_row = row)
            row += 1

        self.set_paper(sheet, start_row - 1)
        self.write_risk_summary(sheet, 1, total_risks, total_initial_scores, start_col + 5, total_residual_scores, start_col + 6)
        logger.info(f"done adding report sheet '{sheet.title}'")

    #
    # bugs
    #
    def bugs_list(self, sheet: openpyxl.worksheet, start_row: int, start_col: int, props: dict = { }) -> None:
        logger.info(f"adding report sheet '{sheet.title}'")

        # bugs, sorted by key
        row = start_row
        for bug in self.jira.sorted_by_fix_version(self.jira.exclude_junk(self.jira.bugs.values(), enforce_versions = False)):
            log_issue(bug)
            col = start_col

            self.write_key_and_summary(sheet, row, col, bug)
            self.write_status(sheet, row, col + 2, bug)
            self.write(sheet, row, col + 3, ', '.join(bug.fix_versions))
            self.write(sheet, row, col + 4, bug.reason_for_deferral)
            self.write(sheet, row, col + 5, bug.risk_level)
            if not bug.reason_for_deferral:
                logger.warn(f"{bug.key} ({','.join(bug.fix_versions)}): unresolved anomaly without reason for deferral")
            row += 1
        if row == start_row:
            self.write(sheet, row, start_col, props['empty'], end_col = start_col + 4)

        self.set_paper(sheet, start_row - 1)
        logger.info(f"done adding report sheet '{sheet.title}'")

    #
    # tests
    #
    def tests_list(self, sheet: openpyxl.worksheet, start_row: int, start_col: int, props: dict = { }) -> None:
        logger.info(f"adding report sheet '{sheet.title}'")

        # tests
        row = start_row
        for report_name, test_report in self.test_reports.reports.items():
            for test in sorted(test_report.test_cases, key = attrgetter('suite', 'name')):
                col = start_col
                self.write(sheet, row, col, report_name)
                self.write(sheet, row, col + 1, test.suite)
                self.write(sheet, row, col + 2, test.name)
                self.write(sheet, row, col + 3, test.time)
                self.write(sheet, row, col + 4, self.passed if test.status else '', format = 'bold')
                row += 1

        self.set_paper(sheet, start_row - 1)
        logger.info(f"done adding report sheet '{sheet.title}'")

    #
    # helper methods
    #

    def range_to_str(self, start_row, start_col, end_row, end_col) -> str:
        end_row = end_row or start_row
        end_col = end_col or start_col
        return f"{openpyxl.utils.cell.get_column_letter(start_col)}{start_row + 1}:{openpyxl.utils.cell.get_column_letter(end_col)}{end_row + 1}"

    def percentage(self, total, count) -> str:
        total = total if total > 0 else 1
        return f'{count} ({round(float(count) / total * 100, 2)}%)'

    def create_format(self, name: str, *formats: List[dict]) -> openpyxl.styles.NamedStyle:
        format = {**self.config['formats']['base']}
        for fmt in formats:
            format.update(fmt)
        if 'bg_color' in format:
            bg_color = openpyxl.styles.colors.Color(rgb = format['bg_color'])
            fill = openpyxl.styles.fills.PatternFill(fill_type = 'solid', start_color = bg_color)
        else:
            fill = openpyxl.styles.fills.PatternFill(fill_type = None)
        color = openpyxl.styles.colors.Color(rgb = format['font_color'])
        font = openpyxl.styles.fonts.Font(name = format['font_name'], color = color, bold = format.get('bold'), size = format.get('font_size'), underline = format.get('underline'))
        align = openpyxl.styles.alignment.Alignment(horizontal = format['align'], vertical = format['valign'], wrap_text = format.get('text_wrap'))
        border_style = openpyxl.styles.borders.Side(style = format.get('border'))
        border = openpyxl.styles.borders.Border(left = border_style, right = border_style, top = border_style, bottom = border_style)
        return openpyxl.styles.NamedStyle(name = name, font = font, alignment = align, border = border, fill = fill)

    def set_outline(self, sheet: openpyxl.worksheet, row: int, start_row: int, level: int = 1) -> None:
        if row > start_row:
            sheet.row_dimensions.group(row, hidden = False, outline_level = level)
        return

    def write_tests(self, sheet: openpyxl.worksheet, row: int, col: int, issue) -> Tuple[int, bool]:
        test_row = row
        verified = False
        for test in self.jira.sorted_by_key(issue.tests):
            self.write_key_and_summary(sheet, test_row, col, test)
            verified = verified or test.is_done
            self.set_outline(sheet, test_row, row, 1)
            test_row += 1
        if test_row == row: # there were no Xray tests
            self.write_key(sheet, test_row, col, issue)
            self.write(sheet, test_row, col + 1, self.labels['see_test_strategy'].format(story_key = issue.key))
            self.write_status(sheet, test_row, col + 2, issue)
            verified = verified or issue.is_done
        return ( test_row, verified )

    def write_risks(self, sheet: openpyxl.worksheet, row: int, col: int, issue, filter: List[str]) -> Tuple[int, bool]:
        risk_row = row
        mitigated = False
        if filter:
            logger.info(f"filtering risks by keys: {filter}")
        for risk in self.jira.sorted_by_key(self.jira.exclude_junk(self.jira.filter_by_key(issue.risks, filter), enforce_versions = False)):
            self.write_key_and_summary(sheet, risk_row, col, risk)
            mitigated = mitigated or risk.is_done
            self.set_outline(sheet, risk_row, row, 1)
            risk_row += 1
        return ( risk_row, mitigated )

    def risk_format(self, risk_score: JiraRiskScore) -> str:
        return formats[risk_score]

    def write_risk_summary(self, sheet: openpyxl.worksheet, start_row: int, total_risks: int, initial_risk_scores: dict, initial_risk_column: int, residual_risk_scores: dict, residual_risk_column: int) -> None:
        logger.info(f'total of {total_risks} risks')
        row = start_row
        for key, count in initial_risk_scores.items():
            logger.info(f'total of initial risk {key.name}: {self.percentage(total_risks, count)} risks')
            self.write(sheet, row, initial_risk_column, self.percentage(total_risks, count), format = self.risk_formats[key])
            row += 1
        self.write(sheet, row, initial_risk_column, self.percentage(total_risks, total_risks), format = 'bold')
        row = start_row
        for key, count in residual_risk_scores.items():
            logger.info(f'total of residual risk {key.name}: {self.percentage(total_risks, count)} risks')
            self.write(sheet, row, residual_risk_column, self.percentage(total_risks, count), format = self.risk_formats[key])
            row += 1
        self.write(sheet, row, residual_risk_column, self.percentage(total_risks, total_risks), format = 'bold')

    def merge(self, sheet: openpyxl.worksheet, row: int, col: int, end_row: int = None, end_col: int = None) -> bool:
        end_row = end_row or row
        end_col = end_col or col
        if end_row - row > 0 or end_col - col > 0:
            sheet.merge_cells(start_row = row, start_column = col, end_row = end_row, end_column = end_col)
            return True
        return False

    def write(self, sheet: openpyxl.worksheet, row: int, col: int, value, format: str = 'base', end_row: int = None, end_col: int = None, url: str = None) -> None:
        cell = sheet.cell(row, col)
        cell.value = value
        if url:
            cell.hyperlink = url
        cell.style = format
        self.merge(sheet, row, col, end_row, end_col)

    def write_key_and_summary(self, sheet: openpyxl.worksheet, row: int, col: int, issue, end_row: int = None, end_col: int = None) -> None:
        self.write_key(sheet, row, col, issue, end_row = end_row)
        self.write_html(sheet, row, col + 1, issue.summary, end_row = end_row)
        if issue.is_test:
            self.write_status(sheet, row, col + 2, issue, end_row = end_row)

    def write_key(self, sheet: openpyxl.worksheet, row: int, col: int, issue, end_row: int = None, end_col: int = None) -> None:
        self.write_url(sheet, row, col, issue, end_row = end_row, end_col = end_col)

    def write_url(self, sheet: openpyxl.worksheet, row: int, col: int, issue, end_row: int = None, end_col: int = None) -> None:
        if self.config['links']:
            self.write(sheet, row, col, issue.key, format = 'url', end_row = end_row, end_col = end_col, url = issue.url)
        else:
            self.write(sheet, row, col, issue.key, format = 'base', end_row = end_row, end_col = end_col)

    def write_html(self, sheet: openpyxl.worksheet, row: int, col: int, value, end_row: int = None, end_col: int = None) -> None:
        html = HtmlToExcel().parse(value)
        self.write(sheet, row, col, html.rendered, format = 'summary', end_row = end_row, end_col = end_col)

    def write_status(self, sheet: openpyxl.worksheet, row: int, col: int, issue, end_row: int = None) -> None:
        if issue.is_done:
            self.write(sheet, row, col, self.passed, format = 'bold', end_row = end_row)
        elif issue.is_blocked:
            self.write(sheet, row, col, self.blocked, format = 'bold', end_row = end_row)
        else:
            self.write(sheet, row, col, issue.status, end_row = end_row)

    def write_id(self, sheet: openpyxl.worksheet, row: int, col: int, issue, end_row: int = None, end_col: int = None) -> None:
        self.write(sheet, row, col, str(issue.id), end_row = end_row, end_col = end_col)

    def set_paper(self, sheet: openpyxl.worksheet, header_row: int = None) -> None:
        sheet.page_setup.paperSize = openpyxl.worksheet.worksheet.Worksheet.PAPERSIZE_LEGAL
        sheet.sheet_properties.outlinePr = openpyxl.worksheet.properties.Outline(summaryBelow = False, summaryRight = False)
        sheet.print_area = sheet.dimensions
        if header_row:
            sheet.print_title_rows = f"{header_row}:{header_row}"
        self.set_header_or_footer(sheet.oddHeader, self.config['page'].get('header'))
        self.set_header_or_footer(sheet.oddFooter, self.config['page'].get('footer'))

    def format_text(self, text: str) -> str:
        if text:
            return text.format(timestamp = self.generated_date, build_number = self.build_number, tag = self.tag)
        return ''

    def set_header_or_footer(self, elem, sections) -> None:
        if sections:
            for key, text in sections.items():
                text = self.format_text(text)
                if key == 'left':
                    elem.left.text = text
                elif key == 'center':
                    elem.center.text = text
                elif key == 'right':
                    elem.right.text = text
